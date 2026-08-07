#!/usr/bin/env python3
"""
Extract_Logfiles_Excel.py
=========================
Parses log files and saves results to:
  1. Per-model CSV files
  2. A master CSV (all models combined)
  3. A multi-sheet Excel file (one sheet per model + ALL_MODELS summary)

Important:
  - The script reads prediction length from log filenames such as:
        *_pred336.log
        *_pred338.log
  - If pred_len is 338, it is normalized and saved as 336.
  - Therefore, both 336 and 338 log files appear as Pred Len = 336
    in all saved CSV and Excel outputs.

Log format expected for accuracy metrics:
  MAE:0.650741, MSE:0.429458, RMSE:0.655331, MAPE:0.237397,
  MSPE:0.056813, RSE:8.186686, R2:-66.021820, Adj_R2 -66.028535

Efficiency metrics format (pipe-separated):
  Average training time per epoch | 12.34
  Data loader                     | 0.001
  Forward pass                    | 0.003
  ...
"""

import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIG
# ============================================================
LOGS_ROOT = Path("logs")

OUTPUT_DIR = Path("./ExtractedResults")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_XLSX = OUTPUT_DIR / "AllModels_household_data_1min_Results.xlsx"

DATASET_NAME = "household_data_1min"   # None = all datasets

MODELS = [
    "RNN", "LSTM", "GRU", "ResLSTM", "BiLSTM", "ConvLSTM",
    "DLinear", "Informer", "iTransformer", "TimesNet",
    "PatchTST", "TimeMixer", "SparseTSF", "ModernTCN",
    "HDMixer", "Crossformer", "Real_FITS", "Times2D",
    "S_Mamba", "TimePro", "gpt2", "Bert", "TEMPO"
]


# ============================================================
# REGEX PATTERNS
# ============================================================
PATTERNS = {
    # ---- Accuracy metrics ----
    "test_mae":    r"(?i)\bMAE:([-\d.]+)",
    "test_mse":    r"(?i)\bMSE:([-\d.]+)",
    "test_rmse":   r"(?i)\bRMSE:([-\d.]+)",
    "test_mape":   r"(?i)\bMAPE:([-\d.]+)",
    "test_mspe":   r"(?i)\bMSPE:([-\d.]+)",
    "test_rse":    r"(?i)\bRSE:([-\d.]+)",
    "test_r2":     r"(?i)\bR2:([-\d.]+)",
    "test_adj_r2": r"(?i)\bAdj_R2\s*([-\d.]+)",

    # ---- Efficiency metrics ----
    "train_time_per_epoch": r"Average training time per epoch\s*\|\s*([\d.]+)",
    "gpu_mem_avg":          r"Average GPU memory usage\s*\|\s*([\d.]+)",
    "gpu_peak_allocated":   r"Peak allocated.*?\|\s*([\d.]+)",
    "gpu_peak_reserved":    r"Peak reserved.*?\|\s*([\d.]+)",
    "wall_data_loader":     r"Data loader\s*\|\s*([\d.]+)",
    "wall_forward":         r"Forward pass\s*\|\s*([\d.]+)",
    "wall_backward":        r"Backward\+Optimizer\s*\|\s*([\d.]+)",
    "infer_latency_avg":    r"Avg latency per batch.*?\|\s*([\d.]+)",
    "infer_latency_p50":    r"p50 latency per batch.*?\|\s*([\d.]+)",
    "infer_latency_p95":    r"p95 latency per batch.*?\|\s*([\d.]+)",
    "infer_throughput":     r"Throughput.*?\|\s*([\d.]+)",

    # ---- Dataset size ----
    "test_samples": r"\btest\s+(\d+)",
}


# ============================================================
# OUTPUT COLUMNS
# ============================================================
COLUMNS = [
    "seq_len", "pred_len",
    "test_mae", "test_mse", "test_rmse", "test_mape",
    "test_mspe", "test_rse", "test_r2", "test_adj_r2",
    "train_time_per_epoch", "gpu_mem_avg",
    "gpu_peak_allocated", "gpu_peak_reserved",
    "wall_data_loader", "wall_forward", "wall_backward",
    "infer_latency_avg", "infer_latency_p50", "infer_latency_p95",
    "infer_throughput", "test_samples", "file",
]

HEADER_LABELS = {
    "seq_len":              "Seq Len",
    "pred_len":             "Pred Len",
    "test_mae":             "MAE",
    "test_mse":             "MSE",
    "test_rmse":            "RMSE",
    "test_mape":            "MAPE",
    "test_mspe":            "MSPE",
    "test_rse":             "RSE",
    "test_r2":              "R2",
    "test_adj_r2":          "Adj R2",
    "train_time_per_epoch": "Train Time (s)",
    "gpu_mem_avg":          "GPU Mem Avg (MB)",
    "gpu_peak_allocated":   "Peak Alloc (MB)",
    "gpu_peak_reserved":    "Peak Resv (MB)",
    "wall_data_loader":     "DataLoader (s)",
    "wall_forward":         "Forward (s)",
    "wall_backward":        "Backward (s)",
    "infer_latency_avg":    "Lat Avg (s)",
    "infer_latency_p50":    "Lat p50 (s)",
    "infer_latency_p95":    "Lat p95 (s)",
    "infer_throughput":     "Throughput",
    "test_samples":         "Test Samples",
    "file":                 "Log File",
}


# ============================================================
# HELPERS
# ============================================================
def safe_float(x):
    try:
        return float(x)
    except Exception:
        return None


def safe_int(x):
    try:
        return int(x)
    except Exception:
        return None


def normalize_pred_len(pred_len):
    """
    Normalize prediction length before saving results.

    Some log files are named with pred_len = 338, but for reporting
    and comparison, they should be treated as pred_len = 336.

    Therefore:
      336 -> 336
      338 -> 336
      other values -> unchanged
    """
    if pred_len in (336, 338):
        return 336
    return pred_len


def infer_metadata(file_path: Path, logs_root: Path):
    """
    Infer model, dataset, sequence length, and prediction length
    from the log file path/name.

    Expected path structure:
        logs / MODEL / DATASET / SEQ_LEN / file_predXXX.log

    Example:
        logs/LSTM/hourly_load/96/something_pred338.log

    In saved outputs:
        pred338 is written as pred_len = 336.
    """
    rel = file_path.relative_to(logs_root).parts

    model = rel[0] if len(rel) > 0 else None
    dataset = rel[1] if len(rel) > 1 else None

    seq_len = None
    if len(rel) > 2 and str(rel[2]).isdigit():
        seq_len = safe_int(rel[2])

    name = file_path.name

    # Read prediction length from filename.
    # Supports names like:
    #   model_seq96_pred336.log
    #   model_seq96_pred338.log
    #   pred336_anything.log
    m = re.search(r"_pred(\d+)\.log$", name) or re.search(r"pred(\d+)", name)
    pred_len_raw = safe_int(m.group(1)) if m else None

    # Normalize before saving.
    pred_len = normalize_pred_len(pred_len_raw)

    # Also infer seq_len from filename if not available from folder.
    m2 = re.search(r"_seq(\d+)_", name)
    if seq_len is None and m2:
        seq_len = safe_int(m2.group(1))

    return {
        "file": str(file_path),
        "model": model,
        "dataset": dataset,
        "seq_len": seq_len,
        "pred_len": pred_len,
    }


def parse_log(file_path: Path, logs_root: Path):
    text = file_path.read_text(encoding="utf-8", errors="ignore")
    rec = infer_metadata(file_path, logs_root)

    for key, pat in PATTERNS.items():
        m = re.search(pat, text)
        if m:
            if key == "test_samples":
                rec[key] = safe_int(m.group(1))
            else:
                rec[key] = safe_float(m.group(1))
        else:
            rec[key] = None

    return rec


def collect_model(model_name: str, dataset_name, logs_root: Path):
    base = logs_root / model_name / dataset_name if dataset_name else logs_root / model_name

    if not base.exists():
        print(f"  ⚠️  No folder: {base}")
        return pd.DataFrame()

    rows = []

    for root, dirs, files in os.walk(base):
        dirs[:] = [
            d for d in dirs
            if not d.startswith(".") and d != ".ipynb_checkpoints"
        ]

        for f in files:
            if f.startswith("."):
                continue
            if not f.endswith(".log"):
                continue
            if "checkpoint" in f.lower():
                continue

            path = Path(root) / f

            try:
                rec = parse_log(path, logs_root)

                if dataset_name and rec.get("dataset") != dataset_name:
                    continue

                rows.append(rec)

            except Exception as e:
                print(f"  ❌ {path}: {e}")

    if not rows:
        print(f"  ⚠️  No logs for {model_name}")
        return pd.DataFrame()

    df = (
        pd.DataFrame(rows)
        .sort_values(["seq_len", "pred_len", "file"])
        .reset_index(drop=True)
    )

    print(f"  ✅ {model_name}: {len(df)} rows")
    return df


# ============================================================
# EXCEL STYLING
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)

ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
NORMAL_FILL = PatternFill("solid", fgColor="FFFFFF")

CELL_FONT = Font(name="Arial", size=9)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT = {
    "test_mae": "0.000000",
    "test_mse": "0.000000",
    "test_rmse": "0.000000",
    "test_mape": "0.000000",
    "test_mspe": "0.000000",
    "test_rse": "0.000000",
    "test_r2": "0.000000",
    "test_adj_r2": "0.000000",

    "train_time_per_epoch": "0.0000",
    "gpu_mem_avg": "0.00",
    "gpu_peak_allocated": "0.00",
    "gpu_peak_reserved": "0.00",

    "wall_data_loader": "0.000000",
    "wall_forward": "0.000000",
    "wall_backward": "0.000000",

    "infer_latency_avg": "0.000000",
    "infer_latency_p50": "0.000000",
    "infer_latency_p95": "0.000000",
    "infer_throughput": "0.00",

    "seq_len": "0",
    "pred_len": "0",
    "test_samples": "0",
}

COL_WIDTH = {
    "file": 60,
    "seq_len": 9,
    "pred_len": 9,
    "test_samples": 12,
    "model": 15,
}


def write_sheet(ws, df: pd.DataFrame):
    cols = [c for c in COLUMNS if c in df.columns]

    # Include model column if this is the ALL_MODELS sheet.
    if "model" in df.columns and "model" not in cols:
        cols = ["model"] + cols

    df = df[cols].copy()

    # Header
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=HEADER_LABELS.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        fill = ALT_FILL if ri % 2 == 0 else NORMAL_FILL

        for ci, col in enumerate(cols, start=1):
            val = row[col]

            if not isinstance(val, str) and pd.isna(val):
                val = None

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = CELL_FONT
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = LEFT if col == "file" else CENTER

            if col in NUM_FMT and val is not None:
                cell.number_format = NUM_FMT[col]

    # Column widths
    for ci, col in enumerate(cols, start=1):
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = COL_WIDTH.get(col, 15)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.row_dimensions[1].height = 28

    for ri in range(2, len(df) + 2):
        ws.row_dimensions[ri].height = 16


def write_empty_sheet(ws):
    for ci, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=HEADER_LABELS.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for ci, col in enumerate(COLUMNS, start=1):
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = COL_WIDTH.get(col, 15)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.row_dimensions[1].height = 28


# ============================================================
# MAIN
# ============================================================
def main():
    print(f"📂 Logs root : {LOGS_ROOT}")
    print(f"📌 Dataset   : {DATASET_NAME or 'ALL'}")
    print("📌 Pred Len normalization: 338 -> 336\n")

    wb = Workbook()
    wb.remove(wb.active)

    all_frames = []

    for model in MODELS:
        print(f"Processing {model}...")

        df = collect_model(model, DATASET_NAME, LOGS_ROOT)

        ws = wb.create_sheet(title=model)

        if not df.empty:
            # Safety normalization again before writing/saving.
            df["pred_len"] = df["pred_len"].apply(normalize_pred_len)

            write_sheet(ws, df)

            df["model"] = model
            all_frames.append(df)

            # Save per-model CSV.
            csv_path = OUTPUT_DIR / f"{model}_{DATASET_NAME}_results.csv"
            df.to_csv(csv_path, index=False)

        else:
            write_empty_sheet(ws)

    # ---- ALL_MODELS summary sheet ----
    if all_frames:
        master = pd.concat(all_frames, ignore_index=True)

        # Safety normalization again before master outputs.
        master["pred_len"] = master["pred_len"].apply(normalize_pred_len)

        master = (
            master
            .sort_values(["model", "seq_len", "pred_len", "file"])
            .reset_index(drop=True)
        )

        ws_all = wb.create_sheet(title="ALL_MODELS", index=0)

        all_cols = ["model"] + [c for c in COLUMNS if c in master.columns]
        write_sheet(ws_all, master[all_cols])

        # Save master CSV.
        master_csv = OUTPUT_DIR / f"AllModels_{DATASET_NAME}_Results.csv"
        master.to_csv(master_csv, index=False)

        print(f"\n📊 Master CSV: {master_csv}")

    wb.save(OUTPUT_XLSX)

    print(f"🎉 Excel saved: {OUTPUT_XLSX}")
    print(f"   Sheets: ALL_MODELS + {', '.join(MODELS)}")


if __name__ == "__main__":
    main()